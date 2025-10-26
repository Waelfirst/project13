# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)


class ExportComponentsWizard(models.TransientModel):
    _name = 'export.components.wizard'
    _description = 'Export Components to Excel'

    pricing_id = fields.Many2one(
        'project.product.pricing',
        string='Pricing',
        required=True,
        readonly=True
    )
    include_specifications = fields.Boolean(
        string='Include Specifications',
        default=True,
        help='Include component specifications in Additional Code column'
    )
    include_bom_data = fields.Boolean(
        string='Include BOM Data',
        default=True,
        help='Include BOM materials and operations in separate sheets'
    )
    
    def action_export(self):
        """Export components to Excel with specifications"""
        self.ensure_one()
        
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError(_('Please install openpyxl library: pip install openpyxl'))
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create Components sheet
        self._create_components_sheet(wb)
        
        # Create Specifications sheet if requested
        if self.include_specifications:
            self._create_specifications_sheet(wb)
        
        # Create BOM sheets if requested
        if self.include_bom_data:
            self._create_bom_materials_sheet(wb)
            self._create_bom_operations_sheet(wb)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create attachment
        file_data = base64.b64encode(output.read())
        filename = 'Components_%s.xlsx' % self.pricing_id.name.replace('/', '_')
        
        attachment = self.env['ir.attachment'].create({
            'name': filename,
            'datas': file_data,
            'res_model': self._name,
            'res_id': self.id,
            'type': 'binary',
        })
        
        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/%s?download=true' % attachment.id,
            'target': 'new',
        }
    
    def _create_components_sheet(self, wb):
        """Create Components sheet with additional_code field"""
        ws = wb.create_sheet('Components')
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4CAF50', end_color='4CAF50', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Component Name',
            'Quantity',
            'Unit',
            'Weight',
            'Cost Price',
            'Total Cost',
            'Additional Code',
            'BOM Code'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 30  # Component Name
        ws.column_dimensions['B'].width = 12  # Quantity
        ws.column_dimensions['C'].width = 10  # Unit
        ws.column_dimensions['D'].width = 12  # Weight
        ws.column_dimensions['E'].width = 12  # Cost Price
        ws.column_dimensions['F'].width = 12  # Total Cost
        ws.column_dimensions['G'].width = 50  # Additional Code
        ws.column_dimensions['H'].width = 15  # BOM Code
        
        # Data
        row_num = 2
        for component in self.pricing_id.component_line_ids:
            ws.cell(row=row_num, column=1).value = component.component_id.name
            ws.cell(row=row_num, column=1).border = border
            
            ws.cell(row=row_num, column=2).value = component.quantity
            ws.cell(row=row_num, column=2).border = border
            ws.cell(row=row_num, column=2).number_format = '0.00'
            
            ws.cell(row=row_num, column=3).value = component.uom_id.name if component.uom_id else ''
            ws.cell(row=row_num, column=3).border = border
            
            ws.cell(row=row_num, column=4).value = component.weight
            ws.cell(row=row_num, column=4).border = border
            ws.cell(row=row_num, column=4).number_format = '0.00'
            
            ws.cell(row=row_num, column=5).value = component.cost_price
            ws.cell(row=row_num, column=5).border = border
            ws.cell(row=row_num, column=5).number_format = '#,##0.00'
            
            ws.cell(row=row_num, column=6).value = component.total_cost
            ws.cell(row=row_num, column=6).border = border
            ws.cell(row=row_num, column=6).number_format = '#,##0.00'
            
            # Additional Code - specifications separated by " - "
            ws.cell(row=row_num, column=7).value = component.additional_code or ''
            ws.cell(row=row_num, column=7).border = border
            ws.cell(row=row_num, column=7).alignment = Alignment(wrap_text=True)
            
            ws.cell(row=row_num, column=8).value = component.bom_id.code if component.bom_id else ''
            ws.cell(row=row_num, column=8).border = border
            
            row_num += 1
    
    def _create_specifications_sheet(self, wb):
        """Create detailed Specifications sheet"""
        ws = wb.create_sheet('Specifications')
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2196F3', end_color='2196F3', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            'Component Name',
            'Specification',
            'Value',
            'Notes'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 30
        ws.column_dimensions['D'].width = 40
        
        # Data
        row_num = 2
        for component in self.pricing_id.component_line_ids:
            if component.specification_ids:
                for spec in component.specification_ids.sorted('sequence'):
                    ws.cell(row=row_num, column=1).value = component.component_id.name
                    ws.cell(row=row_num, column=1).border = border
                    
                    ws.cell(row=row_num, column=2).value = spec.specification_name
                    ws.cell(row=row_num, column=2).border = border
                    
                    ws.cell(row=row_num, column=3).value = spec.value
                    ws.cell(row=row_num, column=3).border = border
                    
                    ws.cell(row=row_num, column=4).value = spec.notes or ''
                    ws.cell(row=row_num, column=4).border = border
                    ws.cell(row=row_num, column=4).alignment = Alignment(wrap_text=True)
                    
                    row_num += 1
    
    def _create_bom_materials_sheet(self, wb):
        """Create BOM Materials sheet"""
        ws = wb.create_sheet('BOM Materials')
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='FF9800', end_color='FF9800', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['BOM Code', 'Material Name', 'Quantity', 'Unit']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 10
        
        # Data
        row_num = 2
        for component in self.pricing_id.component_line_ids:
            if component.bom_id and component.bom_id.bom_line_ids:
                bom_code = component.bom_id.code or f'BOM-{component.bom_id.id}'
                for bom_line in component.bom_id.bom_line_ids:
                    ws.cell(row=row_num, column=1).value = bom_code
                    ws.cell(row=row_num, column=1).border = border
                    
                    ws.cell(row=row_num, column=2).value = bom_line.product_id.name
                    ws.cell(row=row_num, column=2).border = border
                    
                    ws.cell(row=row_num, column=3).value = bom_line.product_qty
                    ws.cell(row=row_num, column=3).border = border
                    ws.cell(row=row_num, column=3).number_format = '0.00'
                    
                    ws.cell(row=row_num, column=4).value = bom_line.product_uom_id.name
                    ws.cell(row=row_num, column=4).border = border
                    
                    row_num += 1
    
    def _create_bom_operations_sheet(self, wb):
        """Create BOM Operations sheet"""
        ws = wb.create_sheet('BOM Operations')
        
        # Define styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='9C27B0', end_color='9C27B0', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = ['BOM Code', 'Operation Name', 'Workcenter', 'Duration (min)']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
        
        # Data
        row_num = 2
        for component in self.pricing_id.component_line_ids:
            if component.bom_id and component.bom_id.routing_id:
                bom_code = component.bom_id.code or f'BOM-{component.bom_id.id}'
                for operation in component.bom_id.routing_id.operation_ids:
                    ws.cell(row=row_num, column=1).value = bom_code
                    ws.cell(row=row_num, column=1).border = border
                    
                    ws.cell(row=row_num, column=2).value = operation.name
                    ws.cell(row=row_num, column=2).border = border
                    
                    ws.cell(row=row_num, column=3).value = operation.workcenter_id.name if operation.workcenter_id else ''
                    ws.cell(row=row_num, column=3).border = border
                    
                    ws.cell(row=row_num, column=4).value = operation.time_cycle_manual
                    ws.cell(row=row_num, column=4).border = border
                    ws.cell(row=row_num, column=4).number_format = '0.00'
                    
                    row_num += 1
