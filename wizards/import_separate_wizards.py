# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    _logger.warning('openpyxl library not found')
    openpyxl = None


class ImportComponentsOnlyWizard(models.TransientModel):
    """معالج استيراد الأجزاء فقط - المرحلة الأولى"""
    _name = 'import.components.only.wizard'
    _description = 'Import Components Only'

    pricing_id = fields.Many2one('project.product.pricing', string='Pricing', required=True)
    excel_file = fields.Binary(string='Excel File', help='Upload Excel file with components')
    filename = fields.Char(string='Filename')

    notes = fields.Text(
        string='Instructions',
        default="""
📦 STEP 1: Import Components/Parts

Required Columns:
- Component Name (اسم الجزء) - Required
- Quantity (الكمية) - Required  
- Weight (kg) (الوزن) - Optional
- Cost Price (سعر التكلفة) - Optional

Example Row:
Steel Sheet | 2 | 5.5 | 50.00

✨ Products will be auto-created if not found
        """,
        readonly=True
    )

    def action_download_template(self):
        """تحميل Template Excel للأجزاء"""
        self.ensure_one()

        if not openpyxl:
            raise UserError(_('Please install openpyxl: pip install openpyxl'))

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Components'

            # Define styles
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = '📦 STEP 1: IMPORT COMPONENTS (استيراد الأجزاء)'
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # Instructions
            ws.merge_cells('A2:D2')
            ws[
                'A2'] = '✨ Products not found will be created automatically | المنتجات غير الموجودة سيتم إنشاؤها تلقائياً'
            ws['A2'].font = Font(size=10, italic=True, color='7F7F7F')
            ws['A2'].fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.row_dimensions[2].height = 30

            # Headers
            headers = ['Component Name *\n(اسم الجزء)', 'Quantity *\n(الكمية)', 'Weight (kg)\n(الوزن)',
                       'Cost Price\n(سعر التكلفة)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws.row_dimensions[3].height = 40

            # Example data
            examples = [
                ['Steel Sheet AISI 304', 2, 5.5, 50.00],
                ['Plastic Housing ABS', 1, 0.8, 25.00],
                ['Aluminum Profile', 3, 2.1, 35.00],
                ['Screws M6x20', 10, 0.05, 0.50],
                ['Electronic Board PCB', 1, 0.3, 120.00],
            ]

            for row_idx, example in enumerate(examples, 4):
                for col_idx, value in enumerate(example, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.fill = example_fill
                    cell.border = border
                    if col_idx in [2, 3, 4]:
                        cell.alignment = Alignment(horizontal='right')

            # Column widths
            ws.column_dimensions['A'].width = 35
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15

            # Save
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            excel_data = base64.b64encode(output.read())
            filename = 'Step1_Components_Template.xlsx'

            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': excel_data,
                'res_model': self._name,
                'res_id': self.id,
                'type': 'binary',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'new',
            }

        except Exception as e:
            raise UserError(_('Error creating template: %s') % str(e))

    def action_import(self):
        """استيراد الأجزاء"""
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload Excel file!'))

        if not openpyxl:
            raise UserError(_('Please install openpyxl library'))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active

            components_created = 0
            products_created = 0

            # Read data starting from row 4 (after title, instructions, headers)
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row[0]:
                    continue

                component_name = str(row[0]).strip()
                if not component_name or 'Steel Sheet' in component_name:  # Skip example
                    continue

                quantity = float(row[1]) if row[1] else 1.0
                weight = float(row[2]) if len(row) > 2 and row[2] else 0.0
                cost_price = float(row[3]) if len(row) > 3 and row[3] else 0.0

                # Find or create product
                product = self.env['product.product'].search([
                    '|', ('name', '=', component_name),
                    ('default_code', '=', component_name)
                ], limit=1)

                if not product:
                    product = self.env['product.product'].create({
                        'name': component_name,
                        'type': 'product',
                        'standard_price': cost_price,
                        'list_price': cost_price * 1.3,
                    })
                    products_created += 1

                # Create component line
                self.env['project.product.component'].create({
                    'pricing_id': self.pricing_id.id,
                    'component_id': product.id,
                    'quantity': quantity,
                    'weight': weight,
                    'cost_price': cost_price,
                })
                components_created += 1

            message = _('✅ Success!\n%s components imported\n%s products created') % (
                components_created, products_created
            )

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Successful'),
                    'message': message,
                    'type': 'success',
                    'sticky': True,
                }
            }

        except Exception as e:
            raise UserError(_('Import error: %s') % str(e))


class ImportBOMMaterialsWizard(models.TransientModel):
    """معالج استيراد مواد BOM - المرحلة الثانية"""
    _name = 'import.bom.materials.wizard'
    _description = 'Import BOM Materials'

    pricing_id = fields.Many2one('project.product.pricing', string='Pricing', required=True)
    excel_file = fields.Binary(string='Excel File', help='Upload Excel file with BOM materials')
    filename = fields.Char(string='Filename')

    notes = fields.Text(
        string='Instructions',
        default="""
📦 STEP 2: Import BOM Materials

Required Columns:
- Component Name (اسم الجزء) - Must match Step 1
- Material Name (اسم المادة الخام) - Required
- Quantity (الكمية) - Required
- Unit (الوحدة) - Optional (kg, pcs, liter)

Example Row:
Steel Sheet | Steel Raw Material | 6 | kg

✨ Materials will be auto-created as products
✨ BOMs will be created for each component
        """,
        readonly=True
    )

    def action_download_template(self):
        """تحميل Template Excel لمواد BOM"""
        self.ensure_one()

        if not openpyxl:
            raise UserError(_('Please install openpyxl'))

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'BOM Materials'

            # Styles
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = '📦 STEP 2: IMPORT BOM MATERIALS (استيراد مواد قوائم المكونات)'
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # Instructions
            ws.merge_cells('A2:D2')
            ws['A2'] = '⚠️ Component names must match Step 1 | أسماء الأجزاء يجب أن تطابق المرحلة الأولى'
            ws['A2'].font = Font(size=10, italic=True, color='E67E22')
            ws['A2'].fill = PatternFill(start_color='FEF5E7', end_color='FEF5E7', fill_type='solid')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 30

            # Headers
            headers = ['Component Name *\n(اسم الجزء)', 'Material Name *\n(اسم المادة)', 'Quantity *\n(الكمية)',
                       'Unit\n(الوحدة)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws.row_dimensions[3].height = 40

            # Examples
            examples = [
                ['Steel Sheet AISI 304', 'Steel Raw Material Grade A', 6, 'kg'],
                ['Steel Sheet AISI 304', 'Coating Material Silver', 0.5, 'kg'],
                ['Plastic Housing ABS', 'Plastic Pellets ABS', 1.2, 'kg'],
                ['Plastic Housing ABS', 'Paint White RAL9003', 0.15, 'liter'],
                ['Aluminum Profile', 'Aluminum Extrusion 6063', 2.5, 'kg'],
            ]

            for row_idx, example in enumerate(examples, 4):
                for col_idx, value in enumerate(example, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.fill = example_fill
                    cell.border = border

            # Column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 35
            ws.column_dimensions['C'].width = 12
            ws.column_dimensions['D'].width = 12

            # Save
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            excel_data = base64.b64encode(output.read())
            filename = 'Step2_BOM_Materials_Template.xlsx'

            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': excel_data,
                'res_model': self._name,
                'res_id': self.id,
                'type': 'binary',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'new',
            }

        except Exception as e:
            raise UserError(_('Error creating template: %s') % str(e))

    def action_import(self):
        """استيراد مواد BOM"""
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload Excel file!'))

        if not openpyxl:
            raise UserError(_('Please install openpyxl'))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active

            # Group materials by component
            component_materials = {}

            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row[0]:
                    continue

                component_name = str(row[0]).strip()
                material_name = str(row[1]).strip() if row[1] else None

                if not component_name or not material_name or 'Steel Sheet' in component_name:
                    continue

                quantity = float(row[2]) if row[2] else 1.0
                unit = str(row[3]).strip() if len(row) > 3 and row[3] else 'Unit(s)'

                if component_name not in component_materials:
                    component_materials[component_name] = []

                component_materials[component_name].append({
                    'material': material_name,
                    'quantity': quantity,
                    'unit': unit,
                })

            boms_created = 0
            materials_created = 0

            for component_name, materials in component_materials.items():
                # Find component
                component_line = self.env['project.product.component'].search([
                    ('pricing_id', '=', self.pricing_id.id),
                    ('component_id.name', '=', component_name)
                ], limit=1)

                if not component_line:
                    _logger.warning('Component not found: %s', component_name)
                    continue

                # Create BOM
                bom_lines = []
                for mat in materials:
                    # Find or create material
                    material = self.env['product.product'].search([
                        '|', ('name', '=', mat['material']),
                        ('default_code', '=', mat['material'])
                    ], limit=1)

                    if not material:
                        material = self.env['product.product'].create({
                            'name': mat['material'],
                            'type': 'product',
                        })
                        materials_created += 1

                    bom_lines.append((0, 0, {
                        'product_id': material.id,
                        'product_qty': mat['quantity'],
                    }))

                if bom_lines:
                    # Check if BOM exists
                    bom = self.env['mrp.bom'].search([
                        ('product_id', '=', component_line.component_id.id)
                    ], limit=1)

                    if bom:
                        bom.bom_line_ids.unlink()
                        bom.write({'bom_line_ids': bom_lines})
                    else:
                        bom = self.env['mrp.bom'].create({
                            'product_id': component_line.component_id.id,
                            'product_tmpl_id': component_line.component_id.product_tmpl_id.id,
                            'product_qty': 1.0,
                            'type': 'normal',
                            'bom_line_ids': bom_lines,
                        })
                        boms_created += 1

                    component_line.bom_id = bom.id

            message = _('✅ Success!\n%s BOMs created/updated\n%s materials created') % (
                boms_created, materials_created
            )

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Successful'),
                    'message': message,
                    'type': 'success',
                    'sticky': True,
                }
            }

        except Exception as e:
            raise UserError(_('Import error: %s') % str(e))


class ImportBOMOperationsWizard(models.TransientModel):
    """معالج استيراد عمليات BOM - المرحلة الثالثة"""
    _name = 'import.bom.operations.wizard'
    _description = 'Import BOM Operations'

    pricing_id = fields.Many2one('project.product.pricing', string='Pricing', required=True)
    excel_file = fields.Binary(string='Excel File', help='Upload Excel file with BOM operations')
    filename = fields.Char(string='Filename')

    notes = fields.Text(
        string='Instructions',
        default="""
⚙️ STEP 3: Import BOM Operations

Required Columns:
- Component Name (اسم الجزء) - Must match Step 1
- Operation Name (اسم العملية) - Required
- Workcenter (مركز العمل) - Required
- Duration (minutes) (المدة بالدقائق) - Required

Example Row:
Steel Sheet | Cutting | CNC Machine | 15

✨ Workcenters will be auto-created
✨ Operations will be added to BOM routing
        """,
        readonly=True
    )

    def action_download_template(self):
        """تحميل Template Excel لعمليات BOM"""
        self.ensure_one()

        if not openpyxl:
            raise UserError(_('Please install openpyxl'))

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'BOM Operations'

            # Styles
            header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:D1')
            ws['A1'] = '⚙️ STEP 3: IMPORT BOM OPERATIONS (استيراد عمليات قوائم المكونات)'
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # Instructions
            ws.merge_cells('A2:D2')
            ws['A2'] = '⚠️ Component names must match Step 1 | أسماء الأجزاء يجب أن تطابق المرحلة الأولى'
            ws['A2'].font = Font(size=10, italic=True, color='E67E22')
            ws['A2'].fill = PatternFill(start_color='FEF5E7', end_color='FEF5E7', fill_type='solid')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 30

            # Headers
            headers = ['Component Name *\n(اسم الجزء)', 'Operation Name *\n(اسم العملية)', 'Workcenter *\n(مركز العمل)',
                       'Duration (min) *\n(المدة)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws.row_dimensions[3].height = 40

            # Examples
            examples = [
                ['Steel Sheet AISI 304', 'Cutting', 'CNC Machine Center 1', 15],
                ['Steel Sheet AISI 304', 'Bending', 'Press Machine 200T', 10],
                ['Steel Sheet AISI 304', 'Coating', 'Coating Line A', 30],
                ['Plastic Housing ABS', 'Injection Molding', 'Molding Machine 1', 5],
                ['Plastic Housing ABS', 'Painting', 'Paint Booth 1', 10],
                ['Aluminum Profile', 'Cutting', 'CNC Machine Center 2', 12],
                ['Aluminum Profile', 'Anodizing', 'Anodizing Tank', 45],
            ]

            for row_idx, example in enumerate(examples, 4):
                for col_idx, value in enumerate(example, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.fill = example_fill
                    cell.border = border

            # Column widths
            ws.column_dimensions['A'].width = 30
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15

            # Save
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            excel_data = base64.b64encode(output.read())
            filename = 'Step3_BOM_Operations_Template.xlsx'

            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': excel_data,
                'res_model': self._name,
                'res_id': self.id,
                'type': 'binary',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'new',
            }

        except Exception as e:
            raise UserError(_('Error creating template: %s') % str(e))

    def action_import(self):
        """استيراد عمليات BOM"""
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload Excel file!'))

        if not openpyxl:
            raise UserError(_('Please install openpyxl'))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active

            # Group operations by component
            component_operations = {}

            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row[0]:
                    continue

                component_name = str(row[0]).strip()
                operation_name = str(row[1]).strip() if row[1] else None

                if not component_name or not operation_name or 'Steel Sheet' in component_name:
                    continue

                workcenter = str(row[2]).strip() if len(row) > 2 and row[2] else None
                duration = float(row[3]) if len(row) > 3 and row[3] else 0.0

                if component_name not in component_operations:
                    component_operations[component_name] = []

                component_operations[component_name].append({
                    'name': operation_name,
                    'workcenter': workcenter,
                    'duration': duration,
                })

            routings_created = 0
            workcenters_created = 0

            for component_name, operations in component_operations.items():
                # Find component
                component_line = self.env['project.product.component'].search([
                    ('pricing_id', '=', self.pricing_id.id),
                    ('component_id.name', '=', component_name)
                ], limit=1)

                if not component_line or not component_line.bom_id:
                    _logger.warning('Component or BOM not found: %s', component_name)
                    continue

                # Create operations
                operation_lines = []
                sequence = 10

                for op in operations:
                    workcenter = None
                    if op.get('workcenter'):
                        # Find or create workcenter
                        workcenter = self.env['mrp.workcenter'].search([
                            ('name', '=', op['workcenter'])
                        ], limit=1)

                        if not workcenter:
                            workcenter = self.env['mrp.workcenter'].create({
                                'name': op['workcenter'],
                                'code': op['workcenter'][:10].upper().replace(' ', '_'),
                            })
                            workcenters_created += 1

                    operation_lines.append((0, 0, {
                        'name': op['name'],
                        'workcenter_id': workcenter.id if workcenter else False,
                        'time_cycle_manual': op.get('duration', 0),
                        'sequence': sequence,
                    }))
                    sequence += 10

                if operation_lines:
                    bom = component_line.bom_id

                    if bom.routing_id:
                        bom.routing_id.operation_ids.unlink()
                        bom.routing_id.write({'operation_ids': operation_lines})
                    else:
                        routing = self.env['mrp.routing'].create({
                            'name': f'{bom.product_id.name} Routing',
                            'operation_ids': operation_lines,
                        })
                        bom.routing_id = routing.id
                        routings_created += 1

            message = _('✅ Success!\n%s routings created\n%s workcenters created') % (
                routings_created, workcenters_created
            )

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Successful'),
                    'message': message,
                    'type': 'success',
                    'sticky': True,
                }
            }

        except Exception as e:
            raise UserError(_('Import error: %s') % str(e))