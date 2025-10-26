# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    _logger.warning('openpyxl library not found')
    openpyxl = None


class OperationsExcelWizard(models.TransientModel):
    """معالج تصدير واستيراد بيانات العمليات الفعلية"""
    _name = 'operations.excel.wizard'
    _description = 'Operations Excel Import/Export Wizard'

    execution_id = fields.Many2one(
        'work.order.execution',
        string='Work Order Execution',
        required=True,
        readonly=True
    )

    mode = fields.Selection([
        ('export', 'Export'),
        ('import', 'Import'),
    ], string='Mode', default='export', required=True)

    excel_file = fields.Binary(
        string='Excel File',
        help='Upload Excel file with actual data'
    )
    filename = fields.Char(string='Filename')

    include_specifications = fields.Boolean(
        string='Include Specifications',
        default=True,
        help='Include component specifications in export'
    )

    notes = fields.Text(
        string='Instructions',
        compute='_compute_notes',
        readonly=True
    )

    @api.depends('mode')
    def _compute_notes(self):
        for wizard in self:
            if wizard.mode == 'export':
                wizard.notes = """
📤 EXPORT OPERATIONS TO EXCEL

This will export all operations with:
• Production Order, Component, Quantity
• Additional Code and Specifications
• Operation Name, Workcenter
• Expected vs Actual Duration
• Workers and Machines Assigned
• Start and Finish Dates

You can then edit the file and re-import it.
                """
            else:
                wizard.notes = """
📥 IMPORT ACTUAL DATA FROM EXCEL

Required Columns:
• Operation Line ID (don't edit this)
• Actual Duration (minutes)
• Workers Assigned
• Machines Assigned

The file must be previously exported from this system.
                """

    def action_export(self):
        """Export operations to Excel"""
        self.ensure_one()

        if not openpyxl:
            raise UserError(_('Please install openpyxl: pip install openpyxl'))

        # Get all operations
        operations = self.env['work.order.operation.line'].search([
            ('execution_id', '=', self.execution_id.id)
        ], order='production_id, sequence')

        if not operations:
            raise UserError(_('No operations found!'))

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = 'Operations Data'

            # Styles
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

            locked_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            editable_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Title
            ws.merge_cells('A1:S1')
            ws['A1'] = '📋 OPERATIONS ACTUAL DATA - Edit Yellow Columns Only'
            ws['A1'].font = Font(size=14, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[1].height = 30

            # Instructions
            ws.merge_cells('A2:S2')
            ws['A2'] = '⚠️ DO NOT edit gray columns! Only edit YELLOW columns (Actual Duration, Workers, Machines)'
            ws['A2'].font = Font(size=10, italic=True, color='E67E22')
            ws['A2'].fill = PatternFill(start_color='FEF5E7', end_color='FEF5E7', fill_type='solid')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            ws.row_dimensions[2].height = 30

            # Headers
            headers = [
                ('ID', True),  # Hidden/Locked
                ('Project', True),
                ('Product', True),
                ('Production Order', True),
                ('Component', True),
                ('Quantity', True),
                ('Additional Code', True),
                ('Specifications', True),
                ('Operation', True),
                ('Workcenter', True),
                ('State', True),
                ('Qty to Produce', True),
                ('Qty Produced', True),
                ('Progress %', True),
                ('Expected Duration (min)', True),
                ('Actual Duration (min)', False),  # EDITABLE
                ('Workers Assigned', False),  # EDITABLE
                ('Machines Assigned', False),  # EDITABLE
                ('Start Date', True),
            ]

            for col, (header, is_locked) in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            ws.row_dimensions[3].height = 35

            # Data
            row = 4
            for op in operations:
                # Get specifications text
                spec_text = ''
                if self.include_specifications and op.specification_ids:
                    specs = []
                    for spec in op.specification_ids.sorted('sequence'):
                        specs.append('%s: %s' % (spec.specification_name, spec.value))
                    spec_text = ' | '.join(specs)

                data = [
                    op.id,  # ID for import
                    op.project_id.name if op.project_id else '',
                    op.product_id.name if op.product_id else '',
                    op.production_id.name if op.production_id else '',
                    op.component_id.name if op.component_id else '',
                    op.execution_line_id.quantity if op.execution_line_id else 0,
                    op.additional_code or '',
                    spec_text,
                    op.name or '',
                    op.workcenter_id.name if op.workcenter_id else '',
                    dict(self.env['mrp.workorder']._fields['state'].selection).get(op.state, op.state or 'pending'),
                    op.qty_production or 0,
                    op.qty_produced or 0,
                    op.progress_percentage or 0,
                    op.duration_expected or 0,
                    op.actual_duration or 0,  # EDITABLE
                    op.workers_assigned or 0,  # EDITABLE
                    op.machines_assigned or 0,  # EDITABLE
                    op.date_start.strftime('%Y-%m-%d %H:%M') if op.date_start else '',
                ]

                for col, (value, (header, is_locked)) in enumerate(zip(data, headers), 1):
                    cell = ws.cell(row=row, column=col)
                    cell.value = value
                    cell.border = border

                    # Color code: gray for locked, yellow for editable
                    if is_locked:
                        cell.fill = locked_fill
                    else:
                        cell.fill = editable_fill

                    # Alignment
                    if col in [6, 12, 13, 14, 15, 16, 17, 18]:  # Numeric columns
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')

                row += 1

            # Column widths
            ws.column_dimensions['A'].width = 10  # ID
            ws.column_dimensions['B'].width = 25  # Project
            ws.column_dimensions['C'].width = 25  # Product
            ws.column_dimensions['D'].width = 20  # Production Order
            ws.column_dimensions['E'].width = 30  # Component
            ws.column_dimensions['F'].width = 10  # Quantity
            ws.column_dimensions['G'].width = 35  # Additional Code
            ws.column_dimensions['H'].width = 40  # Specifications
            ws.column_dimensions['I'].width = 25  # Operation
            ws.column_dimensions['J'].width = 20  # Workcenter
            ws.column_dimensions['K'].width = 15  # State
            ws.column_dimensions['L'].width = 12  # Qty to Produce
            ws.column_dimensions['M'].width = 12  # Qty Produced
            ws.column_dimensions['N'].width = 12  # Progress
            ws.column_dimensions['O'].width = 18  # Expected Duration
            ws.column_dimensions['P'].width = 18  # Actual Duration
            ws.column_dimensions['Q'].width = 15  # Workers
            ws.column_dimensions['R'].width = 15  # Machines
            ws.column_dimensions['S'].width = 18  # Start Date

            # Save
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            excel_data = base64.b64encode(output.read())
            filename = 'Operations_Actual_Data_%s.xlsx' % self.execution_id.name.replace('/', '_')

            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': excel_data,
                'res_model': self._name,
                'res_id': self.id,
                'type': 'binary',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'new',
            }

        except Exception as e:
            _logger.error('Error exporting operations: %s', str(e))
            raise UserError(_('Error exporting operations: %s') % str(e))

    def action_import(self):
        """Import actual data from Excel"""
        self.ensure_one()

        if not self.excel_file:
            raise UserError(_('Please upload Excel file!'))

        if not openpyxl:
            raise UserError(_('Please install openpyxl library'))

        try:
            file_data = base64.b64decode(self.excel_file)
            wb = load_workbook(io.BytesIO(file_data), data_only=True)
            ws = wb.active

            updates_count = 0
            errors = []

            # Read data starting from row 4 (after title, instructions, headers)
            for row in ws.iter_rows(min_row=4, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue

                try:
                    operation_id = int(row[0])
                    actual_duration = float(row[15]) if row[15] else 0.0
                    workers_assigned = int(row[16]) if row[16] else 0
                    machines_assigned = int(row[17]) if row[17] else 0

                    # Find operation
                    operation = self.env['work.order.operation.line'].browse(operation_id)

                    if operation.exists():
                        operation.write({
                            'actual_duration': actual_duration,
                            'workers_assigned': workers_assigned,
                            'machines_assigned': machines_assigned,
                        })
                        updates_count += 1
                    else:
                        errors.append(_('Operation ID %s not found') % operation_id)

                except Exception as e:
                    errors.append(_('Row error: %s') % str(e))
                    continue

            # Show results
            if errors:
                error_msg = '\n'.join(errors[:10])  # Show first 10 errors
                if len(errors) > 10:
                    error_msg += _('\n... and %s more errors') % (len(errors) - 10)

                raise UserError(_(
                    'Import completed with errors:\n\n'
                    'Updated: %s operations\n'
                    'Errors: %s\n\n'
                    'Error details:\n%s'
                ) % (updates_count, len(errors), error_msg))

            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Successful!'),
                    'message': _('%s operations updated successfully!') % updates_count,
                    'type': 'success',
                    'sticky': True,
                }
            }

        except UserError:
            raise
        except Exception as e:
            _logger.error('Import error: %s', str(e))
            raise UserError(_('Error importing Excel file: %s') % str(e))