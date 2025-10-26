# -*- coding: utf-8 -*-

from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError
import base64
import io
import logging

_logger = logging.getLogger(__name__)

try:
    import xlrd
    from xlrd import open_workbook
except ImportError:
    _logger.warning('xlrd library not found, Excel import will not work')
    xlrd = None

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    _logger.warning('openpyxl library not found, Excel import will not work')
    openpyxl = None


class ImportComponentsWizard(models.TransientModel):
    _name = 'import.components.wizard'
    _description = 'Import Components and BOMs from Excel'

    pricing_id = fields.Many2one(
        'project.product.pricing',
        string='Pricing',
        required=True
    )
    product_id = fields.Many2one(
        'product.product',
        string='Product',
        required=True
    )
    excel_file = fields.Binary(
        string='Excel File',
        help='Upload Excel file with components and BOM data'
    )
    filename = fields.Char(string='Filename')
    
    import_type = fields.Selection([
        ('components_only', 'Components Only'),
        ('components_with_bom', 'Components with BOM Data'),
    ], string='Import Type', default='components_with_bom', required=True)
    
    create_missing_products = fields.Boolean(
        string='Auto-Create Missing Products',
        default=True,
        help='Automatically create products that do not exist in the system'
    )
    
    create_missing_workcenters = fields.Boolean(
        string='Auto-Create Missing Workcenters',
        default=True,
        help='Automatically create workcenters that do not exist in the system'
    )
    
    notes = fields.Text(
        string='Instructions',
        default="""
📋 Excel File Format Instructions:

📦 Sheet 1: COMPONENTS (Required)
Columns: Component Name | Quantity | Weight (kg) | Cost Price | BOM Code

📦 Sheet 2: BOM MATERIALS (Optional)
Columns: BOM Code | Material Name | Quantity | Unit

⚙️ Sheet 3: BOM OPERATIONS (Optional)  
Columns: BOM Code | Operation Name | Workcenter | Duration (min) | Workers Needed

💡 Auto-Creation Features:
✅ Products not found will be created automatically
✅ Workcenters not found will be created automatically
✅ Materials not found will be created as products

📝 Important Notes:
• First row is headers - data starts from row 2
• Use example rows as reference (rows with gray background)
• BOM Code links components to their materials and operations
• Duration should be in minutes
• Click "Download Template" to get formatted Excel file
        """,
        readonly=True
    )
    
    def action_download_template(self):
        """Generate professional Excel template with all features"""
        self.ensure_one()
        
        if not openpyxl:
            raise UserError(_('openpyxl library is required. Please install it: pip install openpyxl'))
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.comments import Comment
            
            # Create workbook
            wb = Workbook()
            
            # Define professional styles
            title_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
            
            header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            instruction_font = Font(name='Calibri', size=10, italic=True, color='7F7F7F')
            instruction_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            
            example_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
            
            note_font = Font(name='Calibri', size=9, italic=True, color='E67E22')
            note_fill = PatternFill(start_color='FEF5E7', end_color='FEF5E7', fill_type='solid')
            
            border = Border(
                left=Side(style='thin', color='D0CECE'),
                right=Side(style='thin', color='D0CECE'),
                top=Side(style='thin', color='D0CECE'),
                bottom=Side(style='thin', color='D0CECE')
            )
            
            # =================== SHEET 1: COMPONENTS ===================
            ws1 = wb.active
            ws1.title = 'Components'
            
            # Title row
            ws1.merge_cells('A1:E1')
            ws1['A1'] = '📦 PRODUCT COMPONENTS - Required Sheet'
            ws1['A1'].font = title_font
            ws1['A1'].fill = title_fill
            ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws1.row_dimensions[1].height = 30
            
            # Instructions row
            ws1.merge_cells('A2:E2')
            ws1['A2'] = ('✨ AUTO-CREATE: Products not found will be created automatically. '
                         'Use product name or internal reference code.')
            ws1['A2'].font = instruction_font
            ws1['A2'].fill = instruction_fill
            ws1['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws1.row_dimensions[2].height = 35
            
            # Important note
            ws1.merge_cells('A3:E3')
            ws1['A3'] = '⚠️ NOTE: BOM Code links this component to materials and operations in other sheets'
            ws1['A3'].font = note_font
            ws1['A3'].fill = note_fill
            ws1['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws1.row_dimensions[3].height = 25
            
            # Headers row
            headers1 = ['Component Name *', 'Quantity *', 'Weight (kg)', 'Cost Price', 'BOM Code']
            for col, header in enumerate(headers1, start=1):
                cell = ws1.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws1.row_dimensions[4].height = 35
            
            # Example data rows
            examples1 = [
                ['Steel Sheet AISI 304', 2, 5.5, 50.00, 'BOM-001'],
                ['Plastic Housing ABS', 1, 0.8, 25.00, 'BOM-002'],
                ['Aluminum Frame Profile', 3, 2.1, 35.00, 'BOM-003'],
                ['Screws M6x20 (no BOM)', 10, 0.05, 0.50, ''],
            ]
            
            for row_idx, example in enumerate(examples1, start=5):
                for col_idx, value in enumerate(example, start=1):
                    cell = ws1.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.fill = example_fill
                    cell.border = border
                    if col_idx in [2, 3, 4]:  # Numeric columns
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Add comments/notes to header cells
            ws1['A4'].comment = Comment('Required field. Product will be auto-created if not found.', 'System')
            ws1['B4'].comment = Comment('Required. Enter numeric quantity needed.', 'System')
            ws1['E4'].comment = Comment('Optional. Used to link with BOM Materials and Operations sheets.', 'System')
            
            # Column widths
            ws1.column_dimensions['A'].width = 35
            ws1.column_dimensions['B'].width = 12
            ws1.column_dimensions['C'].width = 15
            ws1.column_dimensions['D'].width = 15
            ws1.column_dimensions['E'].width = 18
            
            # Add empty input rows with borders
            for row in range(9, 30):
                for col in range(1, 6):
                    cell = ws1.cell(row=row, column=col)
                    cell.border = border
            
            # =================== SHEET 2: BOM MATERIALS ===================
            ws2 = wb.create_sheet('BOM Materials')
            
            # Title
            ws2.merge_cells('A1:D1')
            ws2['A1'] = '📦 BOM MATERIALS - Raw Materials for Each Component'
            ws2['A1'].font = title_font
            ws2['A1'].fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
            ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[1].height = 30
            
            # Instructions
            ws2.merge_cells('A2:D2')
            ws2['A2'] = ('✨ AUTO-CREATE: Materials (products) not found will be created automatically. '
                         'Use BOM Code to link with components.')
            ws2['A2'].font = instruction_font
            ws2['A2'].fill = instruction_fill
            ws2['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws2.row_dimensions[2].height = 35
            
            # Important note
            ws2.merge_cells('A3:D3')
            ws2['A3'] = '⚠️ BOM Code must match the BOM Code from Components sheet'
            ws2['A3'].font = note_font
            ws2['A3'].fill = note_fill
            ws2['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws2.row_dimensions[3].height = 25
            
            # Headers
            headers2 = ['BOM Code *', 'Material Name *', 'Quantity *', 'Unit']
            for col, header in enumerate(headers2, start=1):
                cell = ws2.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws2.row_dimensions[4].height = 35
            
            # Example data
            examples2 = [
                ['BOM-001', 'Steel Raw Material Grade A', 6, 'kg'],
                ['BOM-001', 'Coating Material Silver', 0.5, 'kg'],
                ['BOM-002', 'Plastic Pellets ABS Grade', 1.2, 'kg'],
                ['BOM-002', 'Paint White Gloss RAL9003', 0.15, 'liter'],
                ['BOM-002', 'UV Stabilizer Additive', 0.05, 'kg'],
                ['BOM-003', 'Aluminum Extrusion 6063', 2.5, 'kg'],
                ['BOM-003', 'Anodizing Chemical Bath', 0.3, 'liter'],
            ]
            
            for row_idx, example in enumerate(examples2, start=5):
                for col_idx, value in enumerate(example, start=1):
                    cell = ws2.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.fill = example_fill
                    cell.border = border
                    if col_idx == 3:  # Quantity column
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Comments
            ws2['A4'].comment = Comment('Must match BOM Code from Components sheet', 'System')
            ws2['B4'].comment = Comment('Material will be auto-created if not found', 'System')
            ws2['C4'].comment = Comment('Numeric quantity required per unit', 'System')
            ws2['D4'].comment = Comment('Unit of measure (kg, liter, pcs, meters, etc.)', 'System')
            
            # Column widths
            ws2.column_dimensions['A'].width = 18
            ws2.column_dimensions['B'].width = 38
            ws2.column_dimensions['C'].width = 12
            ws2.column_dimensions['D'].width = 12
            
            # Add empty rows
            for row in range(12, 30):
                for col in range(1, 5):
                    cell = ws2.cell(row=row, column=col)
                    cell.border = border
            
            # =================== SHEET 3: BOM OPERATIONS ===================
            ws3 = wb.create_sheet('BOM Operations')
            
            # Title
            ws3.merge_cells('A1:E1')
            ws3['A1'] = '⚙️ BOM OPERATIONS - Manufacturing Operations & Workcenters'
            ws3['A1'].font = title_font
            ws3['A1'].fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
            ws3.row_dimensions[1].height = 30
            
            # Instructions
            ws3.merge_cells('A2:E2')
            ws3['A2'] = ('✨ AUTO-CREATE: Workcenters not found will be created automatically. '
                         'Duration in minutes, workers count is optional.')
            ws3['A2'].font = instruction_font
            ws3['A2'].fill = instruction_fill
            ws3['A2'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            ws3.row_dimensions[2].height = 35
            
            # Important note
            ws3.merge_cells('A3:E3')
            ws3['A3'] = '⚠️ BOM Code must match the BOM Code from Components sheet'
            ws3['A3'].font = note_font
            ws3['A3'].fill = note_fill
            ws3['A3'].alignment = Alignment(horizontal='center', vertical='center')
            ws3.row_dimensions[3].height = 25
            
            # Headers
            headers3 = ['BOM Code *', 'Operation Name *', 'Workcenter *', 'Duration (min) *', 'Workers Needed']
            for col, header in enumerate(headers3, start=1):
                cell = ws3.cell(row=4, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border
            ws3.row_dimensions[4].height = 35
            
            # Example data with workers
            examples3 = [
                ['BOM-001', 'Cutting Steel Sheet', 'CNC Machine Center 1', 15, 1],
                ['BOM-001', 'Bending Operation', 'Press Machine 200T', 10, 2],
                ['BOM-001', 'Coating Application', 'Coating Line A', 30, 1],
                ['BOM-001', 'Quality Inspection', 'QC Station 1', 5, 1],
                ['BOM-002', 'Injection Molding', 'Molding Machine 1', 5, 1],
                ['BOM-002', 'Painting Process', 'Paint Booth 1', 10, 1],
                ['BOM-002', 'Drying Oven', 'Drying Oven 2', 20, 0],
                ['BOM-002', 'Final Assembly', 'Assembly Line A', 8, 2],
                ['BOM-003', 'Aluminum Cutting', 'CNC Machine Center 2', 12, 1],
                ['BOM-003', 'Anodizing Process', 'Anodizing Tank 1', 45, 1],
                ['BOM-003', 'Packaging', 'Pack Station 1', 5, 1],
            ]
            
            for row_idx, example in enumerate(examples3, start=5):
                for col_idx, value in enumerate(example, start=1):
                    cell = ws3.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.fill = example_fill
                    cell.border = border
                    if col_idx in [4, 5]:  # Numeric columns
                        cell.alignment = Alignment(horizontal='right', vertical='center')
                    else:
                        cell.alignment = Alignment(horizontal='left', vertical='center')
            
            # Comments
            ws3['A4'].comment = Comment('Must match BOM Code from Components sheet', 'System')
            ws3['B4'].comment = Comment('Name of the manufacturing operation', 'System')
            ws3['C4'].comment = Comment('Workcenter will be auto-created if not found', 'System')
            ws3['D4'].comment = Comment('Operation duration in minutes', 'System')
            ws3['E4'].comment = Comment('Optional: Number of workers needed for this operation', 'System')
            
            # Column widths
            ws3.column_dimensions['A'].width = 18
            ws3.column_dimensions['B'].width = 30
            ws3.column_dimensions['C'].width = 30
            ws3.column_dimensions['D'].width = 18
            ws3.column_dimensions['E'].width = 18
            
            # Add empty rows
            for row in range(16, 30):
                for col in range(1, 6):
                    cell = ws3.cell(row=row, column=col)
                    cell.border = border
            
            # =================== SAVE WORKBOOK ===================
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            excel_data = base64.b64encode(output.read())
            
            # Create attachment
            filename = 'Component_Import_Template_%s.xlsx' % fields.Date.today().strftime('%Y%m%d')
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'datas': excel_data,
                'res_model': self._name,
                'res_id': self.id,
                'type': 'binary',
            })
            
            return {
                'type': 'ir.actions.act_url',
                'url': '/web/content/%s?download=true' % attachment.id,
                'target': 'new',
            }
            
        except Exception as e:
            _logger.error('Error creating template: %s', str(e))
            raise UserError(_('Error creating template: %s') % str(e))
    
    def action_import(self):
        """Import data from uploaded Excel file with auto-creation"""
        self.ensure_one()
        
        if not self.excel_file:
            raise UserError(_('Please upload an Excel file!'))
        
        if not openpyxl:
            raise UserError(_('openpyxl library not installed. Please install it: pip install openpyxl'))
        
        try:
            # Decode the file
            file_data = base64.b64decode(self.excel_file)
            
            # Load workbook
            workbook = load_workbook(io.BytesIO(file_data), data_only=True)
            
            # Import with auto-creation enabled
            result = self._import_with_openpyxl(workbook)
            
            # Prepare success message
            message_parts = []
            message_parts.append(_('%s components imported successfully!') % result['components_count'])
            
            if result.get('products_created'):
                message_parts.append(_('%s products auto-created') % result['products_created'])
            
            if result.get('workcenters_created'):
                message_parts.append(_('%s workcenters auto-created') % result['workcenters_created'])
            
            if result.get('boms_created'):
                message_parts.append(_('%s BOMs created') % result['boms_created'])
            
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Import Successful!'),
                    'message': '\n'.join(message_parts),
                    'type': 'success',
                    'sticky': True,
                }
            }
            
        except Exception as e:
            _logger.error('Import error: %s', str(e))
            raise UserError(_('Error importing Excel file: %s') % str(e))
    
    def _import_with_openpyxl(self, workbook):
        """Import using openpyxl with auto-creation support"""
        components_data = []
        bom_materials = {}
        bom_operations = {}
        
        stats = {
            'components_count': 0,
            'products_created': 0,
            'workcenters_created': 0,
            'boms_created': 0,
        }
        
        # Read Components Sheet (skip title rows 1-4)
        if 'Components' in workbook.sheetnames:
            ws = workbook['Components']
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]:  # Skip empty rows
                    continue
                
                component_name = str(row[0]).strip()
                if not component_name or component_name.startswith('Example') or component_name.startswith('Screws'):
                    continue
                
                quantity = float(row[1]) if row[1] else 1.0
                weight = float(row[2]) if row[2] else 0.0
                cost_price = float(row[3]) if row[3] else 0.0
                bom_code = row[4] if len(row) > 4 and row[4] else None
                
                components_data.append({
                    'name': component_name,
                    'quantity': quantity,
                    'weight': weight,
                    'cost_price': cost_price,
                    'bom_code': bom_code,
                })
        
        # Read BOM Materials Sheet (skip title rows 1-4)
        if 'BOM Materials' in workbook.sheetnames and self.import_type == 'components_with_bom':
            ws = workbook['BOM Materials']
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]:
                    continue
                
                bom_code = str(row[0]).strip()
                material_name = str(row[1]).strip() if row[1] else None
                
                if not bom_code or not material_name:
                    continue
                
                quantity = float(row[2]) if row[2] else 1.0
                unit = str(row[3]).strip() if len(row) > 3 and row[3] else None
                
                if bom_code not in bom_materials:
                    bom_materials[bom_code] = []
                
                bom_materials[bom_code].append({
                    'material': material_name,
                    'quantity': quantity,
                    'unit': unit,
                })
        
        # Read BOM Operations Sheet (skip title rows 1-4)
        if 'BOM Operations' in workbook.sheetnames and self.import_type == 'components_with_bom':
            ws = workbook['BOM Operations']
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]:
                    continue
                
                bom_code = str(row[0]).strip()
                operation_name = str(row[1]).strip() if row[1] else None
                workcenter_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
                duration = float(row[3]) if len(row) > 3 and row[3] else 0.0
                workers_needed = int(row[4]) if len(row) > 4 and row[4] else 1
                
                if not bom_code or not operation_name:
                    continue
                
                if bom_code not in bom_operations:
                    bom_operations[bom_code] = []
                
                bom_operations[bom_code].append({
                    'name': operation_name,
                    'workcenter': workcenter_name,
                    'duration': duration,
                    'workers_needed': workers_needed,
                })
        
        # Create components and BOMs with auto-creation
        result = self._create_components_and_boms_with_autocreate(
            components_data, 
            bom_materials, 
            bom_operations
        )
        
        return result
    
    def _create_components_and_boms_with_autocreate(self, components_data, bom_materials, bom_operations):
        """Create component lines and BOMs with auto-creation of missing items"""
        stats = {
            'components_count': 0,
            'products_created': 0,
            'workcenters_created': 0,
            'boms_created': 0,
        }
        
        for comp_data in components_data:
            # Find or create product
            product = self._find_or_create_product(comp_data['name'], comp_data['cost_price'])
            if not product:
                continue
            
            if product and not self.env['product.product'].search([('id', '=', product.id)], limit=1):
                stats['products_created'] += 1
            
            # Create component line
            component = self.env['project.product.component'].create({
                'pricing_id': self.pricing_id.id,
                'component_id': product.id,
                'quantity': comp_data['quantity'],
                'weight': comp_data['weight'],
                'cost_price': comp_data['cost_price'],
            })
            
            # Create BOM if requested
            bom_code = comp_data.get('bom_code')
            if bom_code and self.import_type == 'components_with_bom':
                bom_result = self._create_bom_with_autocreate(
                    product,
                    bom_code,
                    bom_materials.get(bom_code, []),
                    bom_operations.get(bom_code, [])
                )
                
                if bom_result['bom']:
                    component.bom_id = bom_result['bom'].id
                    stats['boms_created'] += 1
                    stats['products_created'] += bom_result['materials_created']
                    stats['workcenters_created'] += bom_result['workcenters_created']
            
            stats['components_count'] += 1
        
        return stats
    
    def _find_or_create_product(self, product_name, cost_price=0.0):
        """Find existing product or create new one"""
        # Try to find existing product
        product = self.env['product.product'].search([
            '|', ('name', '=', product_name),
            ('default_code', '=', product_name)
        ], limit=1)
        
        if product:
            return product
        
        # Auto-create if enabled
        if self.create_missing_products:
            try:
                product = self.env['product.product'].create({
                    'name': product_name,
                    'type': 'product',
                    'standard_price': cost_price,
                    'list_price': cost_price * 1.3,  # 30% markup
                    'detailed_type': 'product',
                    'categ_id': self.env.ref('product.product_category_all').id,
                })
                _logger.info('Auto-created product: %s', product_name)
                return product
            except Exception as e:
                _logger.error('Error auto-creating product %s: %s', product_name, str(e))
                return None
        else:
            _logger.warning('Product not found and auto-create disabled: %s', product_name)
            return None
    
    def _find_or_create_workcenter(self, workcenter_name):
        """Find existing workcenter or create new one"""
        # Try to find existing workcenter
        workcenter = self.env['mrp.workcenter'].search([
            ('name', '=', workcenter_name)
        ], limit=1)
        
        if workcenter:
            return workcenter
        
        # Auto-create if enabled
        if self.create_missing_workcenters:
            try:
                workcenter = self.env['mrp.workcenter'].create({
                    'name': workcenter_name,
                    'code': workcenter_name[:10].upper().replace(' ', '_'),
                    'time_efficiency': 100,
                    'time_start': 0,
                    'time_stop': 0,
                })
                _logger.info('Auto-created workcenter: %s', workcenter_name)
                return workcenter
            except Exception as e:
                _logger.error('Error auto-creating workcenter %s: %s', workcenter_name, str(e))
                return None
        else:
            _logger.warning('Workcenter not found and auto-create disabled: %s', workcenter_name)
            return None
    
    def _create_bom_with_autocreate(self, product, bom_code, materials, operations):
        """Create BOM with materials and operations, auto-creating missing items"""
        result = {
            'bom': None,
            'materials_created': 0,
            'workcenters_created': 0,
        }
        
        try:
            # Check if BOM already exists
            existing_bom = self.env['mrp.bom'].search([
                ('product_id', '=', product.id),
                ('code', '=', bom_code)
            ], limit=1)
            
            if existing_bom:
                result['bom'] = existing_bom
                return result
            
            # Create BOM lines (materials)
            bom_lines = []
            for mat in materials:
                material_product = self._find_or_create_product(mat['material'])
                
                if material_product:
                    # Check if this is a newly created product
                    if not self.env['product.product'].search([
                        ('name', '=', mat['material'])
                    ], limit=1):
                        result['materials_created'] += 1
                    
                    bom_lines.append((0, 0, {
                        'product_id': material_product.id,
                        'product_qty': mat['quantity'],
                    }))
            
            # Create routing operations
            routing_lines = []
            for op in operations:
                workcenter = None
                if op.get('workcenter'):
                    # Count before creation
                    wc_exists = self.env['mrp.workcenter'].search([
                        ('name', '=', op['workcenter'])
                    ], limit=1)
                    
                    workcenter = self._find_or_create_workcenter(op['workcenter'])
                    
                    # Count after creation
                    if workcenter and not wc_exists:
                        result['workcenters_created'] += 1
                
                if workcenter or not op.get('workcenter'):
                    operation_vals = {
                        'name': op['name'],
                        'workcenter_id': workcenter.id if workcenter else False,
                        'time_cycle_manual': op.get('duration', 0),
                    }
                    
                    # Add workers if specified
                    if op.get('workers_needed'):
                        operation_vals['workorder_count'] = op['workers_needed']
                    
                    routing_lines.append((0, 0, operation_vals))
            
            # Create BOM
            bom_vals = {
                'product_id': product.id,
                'product_tmpl_id': product.product_tmpl_id.id,
                'product_qty': 1.0,
                'type': 'normal',
                'code': bom_code,
                'bom_line_ids': bom_lines,
            }
            
            # Create routing if operations exist
            if routing_lines:
                routing = self.env['mrp.routing'].create({
                    'name': f'{product.name} - {bom_code}',
                    'operation_ids': routing_lines,
                })
                bom_vals['routing_id'] = routing.id
            
            bom = self.env['mrp.bom'].create(bom_vals)
            result['bom'] = bom
            _logger.info('Created BOM %s for product %s', bom_code, product.name)
            
            return result
            
        except Exception as e:
            _logger.error('Error creating BOM for %s: %s', product.name, str(e))
            return result
